#!/usr/bin/env python3
"""
DGR Performance Dashboard — FY 2025-26
=======================================
Reads all 4 plant DGR Excel files → Fleet overview + per-plant KPIs.

Run:  python dgr_dashboard.py
Open: http://localhost:8051
"""
import base64, io, os, datetime, re, sys
from collections import defaultdict

# Ensure Unicode characters print correctly on Windows CMD (CP1252 terminals)
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

import openpyxl
import dash
from dash import dcc, html, Input, Output, State
import plotly.graph_objects as go

# ── Config ────────────────────────────────────────────────────────────────────
# Data folder: works both locally (data/ subfolder) and on Render
DGR_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
TARIFF     = 2.75      # ₹ per kWh (PPA rate — update to your actual rate)
PR_TARGET  = 75.0      # % benchmark
CO2_FACTOR = 0.000716  # tonne CO₂ per kWh (CEA India 2023)

BLU = "#2563eb"; GRN = "#16a34a"; AMB = "#d97706"
RED = "#dc2626"; PUR = "#7c3aed"; TEL = "#0891b2"
PLANT_COLS = [BLU, GRN, AMB, PUR, TEL, RED]

CHART_CFG = dict(
    plot_bgcolor="#f8fafc", paper_bgcolor="#ffffff",
    font=dict(family="Arial, sans-serif", color="#1e293b"),
    margin=dict(l=50, r=20, t=48, b=50),
    hovermode="x unified",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
)

# ── UI helpers ────────────────────────────────────────────────────────────────

def _card(children, extra=None):
    s = {"backgroundColor":"#fff","borderRadius":"12px","padding":"10px",
         "boxShadow":"0 1px 3px rgba(0,0,0,0.08)"}
    if extra: s.update(extra)
    return html.Div(children, style=s)

def _ef(msg="Upload DGR files or click 'Load DGR Folder'"):
    f = go.Figure()
    f.update_layout(**CHART_CFG)
    f.add_annotation(text=msg, x=0.5, y=0.5, showarrow=False,
                     font=dict(size=14, color="#94a3b8"), xref="paper", yref="paper")
    return f

def _alert(msg, kind="info"):
    styles = {
        "success": ("#f0fdf4","#bbf7d0","#166534"),
        "error":   ("#fef2f2","#fecaca","#dc2626"),
        "info":    ("#eff6ff","#bfdbfe","#1e40af"),
    }
    bg, br, col = styles[kind]
    icon = {"success":"✅","error":"⚠️","info":"ℹ️"}[kind]
    return html.Div([html.Span(f"{icon}  "), html.Span(msg)], style={
        "backgroundColor": bg, "border": f"1px solid {br}",
        "borderRadius": "8px", "padding": "10px 16px",
        "color": col, "marginBottom": "14px", "fontSize": "13px",
    })

def kpi_card(title, value, unit, sub, color, icon):
    return html.Div([
        html.Div([
            html.Span(icon, style={"fontSize": "18px"}),
            html.Span(title, style={"fontSize":"10px","color":"#64748b","fontWeight":"700",
                                    "textTransform":"uppercase","letterSpacing":"0.08em","marginLeft":"6px"}),
        ], style={"display":"flex","alignItems":"center","marginBottom":"8px"}),
        html.Div([
            html.Span(str(value), style={"fontSize":"28px","fontWeight":"800","color":color}),
            html.Span(f" {unit}", style={"fontSize":"12px","color":"#94a3b8","marginLeft":"2px"}),
        ]),
        html.Hr(style={"border":f"1px solid {color}","opacity":"0.2","margin":"8px 0"}),
        html.Div(sub, style={"fontSize":"11px","color":"#94a3b8","lineHeight":"1.5","whiteSpace":"pre-line"}),
    ], style={
        "backgroundColor":"#fff","borderRadius":"12px","padding":"16px 18px",
        "boxShadow":"0 1px 4px rgba(0,0,0,0.09)","borderLeft":f"5px solid {color}",
        "flex":"1","minWidth":"145px",
    })


# ── Parser helpers ────────────────────────────────────────────────────────────

def _plant_name_from_file(fname):
    n = os.path.splitext(fname)[0]            # strip .xlsx extension first
    n = n.replace("DGR_","").replace("_FY25_26","").replace("_FY 25 26","")
    n = re.sub(r'\s*\(\d+\)','', n).strip().replace("_"," ")
    return n

def _td_to_hrs(val):
    if isinstance(val, datetime.timedelta): return val.total_seconds()/3600
    if isinstance(val, datetime.time):      return val.hour + val.minute/60 + val.second/3600
    if isinstance(val, (int,float)):        return float(val)
    return 0.0

def _safe_float(val, default=0.0):
    """Convert to float, returning default for non-numeric values.
    Handles: None, int/float, European decimal commas ('521,28' → 521.28),
             double-dots ('1425..63' → 1425.63), and plain text descriptions.
    """
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    try:
        return float(val)
    except (ValueError, TypeError):
        try:
            s = str(val).strip().replace(',', '.').replace('..', '.')
            return float(s)
        except (ValueError, TypeError):
            return default

def _to_date(dv):
    if isinstance(dv, datetime.datetime): return dv.date()
    if isinstance(dv, datetime.date):     return dv
    return None


# ── DGR Excel Parser ──────────────────────────────────────────────────────────

def parse_dgr(raw: bytes, filename: str):
    """
    Returns (plant_dict, error_str).
    Reads: Budgeted_Data, Data_Entry, Plant_Down_Time Details, Plant Availiblity
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return None, f"{filename}: {e}"

    plant_name = _plant_name_from_file(filename)
    dc_kw      = 0
    records    = []   # from Budgeted_Data
    inv_recs   = []   # from Data_Entry
    faults     = []   # from Plant_Down_Time Details
    avail_recs = []   # from Plant Availiblity

    # ── Budgeted_Data ─────────────────────────────────────────────────────────
    # Cols: 0=PlantName 1=Date 4=Gen(kWh) 5=GTI 6=GHI 7=Temp 8=PR 9=CUF 12=DC
    ws = next((wb[s] for s in ["Budgeted_Data","Budgeted_data"] if s in wb.sheetnames), None)
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 10: continue
            d = _to_date(row[1])
            if not d: continue
            gen = _safe_float(row[4])
            if gen <= 0: continue

            if len(row) > 12 and row[12]: dc_kw = max(dc_kw, _safe_float(row[12]))

            pr_f  = _safe_float(row[8])
            cuf_f = _safe_float(row[9])
            # Stored as decimal (0.82 = 82%) — convert to %
            pr  = pr_f*100  if pr_f  <= 1.5 else pr_f
            cuf = cuf_f*100 if cuf_f <= 1.5 else cuf_f

            records.append({
                "date":    d.isoformat(),
                "gen_kwh": gen,
                "gti":     _safe_float(row[5]),
                "ghi":     _safe_float(row[6]),
                "temp":    _safe_float(row[7]),
                "pr":      round(pr, 2),
                "cuf":     round(cuf, 2),
            })

    # ── Data_Entry (per-inverter) ─────────────────────────────────────────────
    if "Data_Entry" in wb.sheetnames:
        ws = wb["Data_Entry"]
        hdr = None
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            if row and any("Inv_" in str(v or "") for v in row):
                hdr = row; break
        if hdr:
            inv_cols  = [(i, str(v)) for i,v in enumerate(hdr) if v and str(v).startswith("Inv_")]
            net_col   = next((i for i,v in enumerate(hdr) if v and "Net_Gen"  in str(v)), None)
            gti_col   = next((i for i,v in enumerate(hdr) if v and "GTI"      in str(v)), None)
            wx_col    = next((i for i,v in enumerate(hdr) if v and "Weather"  in str(v)), None)
            DATE_COL  = 2

            for row in ws.iter_rows(min_row=2, values_only=True):
                if DATE_COL >= len(row): continue
                d = _to_date(row[DATE_COL])
                if not d: continue
                inv_data = {cn: _safe_float(row[ci]) for ci,cn in inv_cols
                            if ci < len(row) and row[ci] is not None}
                if not inv_data: continue
                rec = {"date": d.isoformat(),
                       "net_gen": _safe_float(row[net_col]) if net_col and net_col<len(row) else 0,
                       "gti":     _safe_float(row[gti_col]) if gti_col and gti_col<len(row) else 0,
                       "weather": str(row[wx_col] or "")    if wx_col  and wx_col <len(row) else ""}
                rec.update(inv_data)
                inv_recs.append(rec)

    # ── Plant_Down_Time Details ───────────────────────────────────────────────
    # Cols: 1=Date 2=Inverter 4=Start 5=Stop 6=TotalDT 7=AffectedDC 12=FaultDesc 13=Reason
    for sn in ["Plant_Down_Time Details","Plant_Down_Time_Details"]:
        if sn not in wb.sheetnames: continue
        for row in wb[sn].iter_rows(min_row=3, values_only=True):
            if len(row) < 7: continue
            d = _to_date(row[1])
            if not d: continue
            dt_hrs = _td_to_hrs(row[6])
            if dt_hrs <= 0: continue
            r13 = row[13] if len(row) > 13 else None
            r12 = row[12] if len(row) > 12 else None
            reason = str(r13 or r12 or "Unknown").strip().upper()[:50] or "Unknown"
            faults.append({
                "date":         d.isoformat(),
                "inverter":     str(row[2] or ""),
                "downtime_hrs": round(dt_hrs, 2),
                "reason":       reason,
                "dc_affected":  _safe_float(row[7]) if len(row) > 7 else 0,
            })
        break

    # ── Plant Availiblity ─────────────────────────────────────────────────────
    # Cols: 0=Date 1=ABT_Export 3=GHI 4=CUF 8=OutageHrs 13=PlantAvailability
    for sn in ["Plant  Availiblity","Plant Availiblity","Plant_Availiblity","Plant Aviliblity"]:
        if sn not in wb.sheetnames: continue
        for row in wb[sn].iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 14: continue
            d = _to_date(row[0])
            if not d: continue
            av = _safe_float(row[13])
            if av == 0.0: continue
            avail_pct = av*100 if av <= 1.5 else av
            avail_recs.append({
                "date":         d.isoformat(),
                "gen_kwh":      _safe_float(row[1] if len(row) > 1 else None),
                "ghi":          _safe_float(row[3] if len(row) > 3 else None),
                "outage_hrs":   _safe_float(row[8] if len(row) > 8 else None),
                "availability": round(avail_pct, 2),
            })
        break

    if not records:
        return None, f"No generation data in {filename} — check Budgeted_Data sheet."

    return {
        "plant_name":  plant_name,
        "filename":    filename,
        "dc_kw":       round(dc_kw),
        "records":     records,
        "inv_recs":    inv_recs,
        "faults":      faults,
        "avail_recs":  avail_recs,
    }, None


def load_dgr_folder():
    if not os.path.isdir(DGR_FOLDER):
        return None, f"Folder not found: {DGR_FOLDER}"
    files = [f for f in os.listdir(DGR_FOLDER) if f.lower().endswith(".xlsx")]
    if not files:
        return None, "No .xlsx files in DGR folder."
    all_plants, errors = {}, []
    for fname in sorted(files):
        try:
            with open(os.path.join(DGR_FOLDER, fname), "rb") as fh:
                raw = fh.read()
            data, err = parse_dgr(raw, fname)
            if err:   errors.append(err)
            else:     all_plants[data["plant_name"]] = data
        except Exception as e:
            errors.append(f"{fname}: {e}")
    if not all_plants:
        return None, "No valid files. " + " | ".join(errors)
    err_str = " | ".join(errors) if errors else None
    return all_plants, err_str


# ── KPIs ─────────────────────────────────────────────────────────────────────

def compute_kpis(plant):
    recs = plant.get("records", [])
    if not recs: return {}
    n    = len(recs)
    tgen = sum(r["gen_kwh"] for r in recs)
    prs  = [r["pr"]  for r in recs if r["pr"]  > 0]
    cufs = [r["cuf"] for r in recs if r["cuf"] > 0]
    avg_pr  = sum(prs)  / len(prs)  if prs  else 0
    avg_cuf = sum(cufs) / len(cufs) if cufs else 0

    avs = plant.get("avail_recs", [])
    if avs:
        a_vals = [r["availability"] for r in avs if r["availability"] > 0]
        avg_av  = sum(a_vals)/len(a_vals) if a_vals else 100
        tot_out = sum(r["outage_hrs"] for r in avs)
    else:
        faults  = plant.get("faults", [])
        tot_out = sum(f["downtime_hrs"] for f in faults)
        avg_av  = max(0, 100 - tot_out/(n*24)*100) if n else 100

    dates = sorted(r["date"] for r in recs)
    grade, gc = (
        ("Excellent", GRN) if avg_pr >= 80 else
        ("Good",      "#22c55e") if avg_pr >= 75 else
        ("Fair",      AMB) if avg_pr >= 65 else ("Poor", RED)
    )
    return {
        "plant_name":    plant["plant_name"],
        "dc_kw":         plant.get("dc_kw", 0),
        "n_days":        n,
        "date_from":     datetime.date.fromisoformat(dates[0]).strftime("%d %b %Y"),
        "date_to":       datetime.date.fromisoformat(dates[-1]).strftime("%d %b %Y"),
        "total_gen_mwh": round(tgen/1000, 2),
        "avg_pr":        round(avg_pr, 1),
        "avg_cuf":       round(avg_cuf, 2),
        "availability":  round(avg_av, 1),
        "revenue_lakh":  round(tgen*TARIFF/1e5, 2),
        "co2_tonnes":    round(tgen*CO2_FACTOR, 2),
        "outage_hrs":    round(tot_out, 1),
        "n_faults":      len(plant.get("faults", [])),
        "grade":         grade,
        "grade_color":   gc,
    }


# ── Charts ────────────────────────────────────────────────────────────────────

def _pc(p): return GRN if p >= PR_TARGET else AMB if p >= 65 else RED

def chart_daily_gen(plant, days=120):
    recs  = plant["records"][-days:]
    dates = [r["date"]      for r in recs]
    gens  = [r["gen_kwh"]/1000 for r in recs]   # MWh
    prs   = [r["pr"]        for r in recs]

    fig = go.Figure()
    fig.add_bar(x=dates, y=gens, name="Daily Generation (MWh)",
                marker_color=[_pc(p) for p in prs], opacity=0.85,
                hovertemplate="%{x}<br><b>%{y:.2f} MWh</b><extra></extra>")
    fig.add_scatter(x=dates, y=prs, name="PR %", yaxis="y2", mode="lines",
                    line=dict(color="#f59e0b", width=2, dash="dot"),
                    hovertemplate="PR: %{y:.1f}%<extra></extra>")
    fig.add_hline(y=PR_TARGET, yref="y2", line_dash="dash", line_color=RED, line_width=1.5,
                  annotation_text=f"PR Target {PR_TARGET}%", annotation_position="right")
    fig.update_layout(
        **CHART_CFG,
        title=f"{plant['plant_name']}  —  Daily Generation & PR",
        xaxis=dict(tickformat="%d %b", tickangle=-40), xaxis_title="Date",
        yaxis=dict(title="Generation (MWh)"),
        yaxis2=dict(title="PR (%)", overlaying="y", side="right",
                    range=[50, 105], showgrid=False),
    )
    return fig

def chart_pr_trend(plant):
    recs  = plant["records"]
    dates = [r["date"] for r in recs]
    prs   = [r["pr"]   for r in recs]

    fig = go.Figure()
    fig.add_hrect(y0=PR_TARGET, y1=105, fillcolor="rgba(22,163,74,0.07)", line_width=0)
    fig.add_scatter(x=dates, y=prs, name="PR %", mode="lines+markers",
                    line=dict(color=BLU, width=2.5),
                    marker=dict(size=5, color=[_pc(p) for p in prs]),
                    hovertemplate="%{x}  PR: %{y:.1f}%<extra></extra>")
    fig.add_hline(y=PR_TARGET, line_dash="dash", line_color=RED, line_width=2,
                  annotation_text=f"Target {PR_TARGET}%", annotation_position="right")
    lo = max(0, min(prs)-8) if prs else 0
    fig.update_layout(**CHART_CFG, title="Performance Ratio Trend",
                      xaxis=dict(tickformat="%d %b", tickangle=-40), xaxis_title="Date",
                      yaxis=dict(title="PR (%)", range=[lo, 100]))
    return fig

def chart_monthly(plant):
    mon = defaultdict(float)
    for r in plant["records"]:
        k = datetime.date.fromisoformat(r["date"]).strftime("%b '%y")
        mon[k] += r["gen_kwh"]
    months  = list(mon)
    gen_mwh = [mon[m]/1000 for m in months]

    fig = go.Figure()
    fig.add_bar(x=months, y=gen_mwh, name="MWh", marker_color=BLU, opacity=0.85,
                hovertemplate="%{x}: %{y:.1f} MWh<extra></extra>")
    fig.add_scatter(x=months, y=gen_mwh, mode="lines+markers", name="Trend",
                    line=dict(color=GRN, width=2.5), marker=dict(size=8), hoverinfo="skip")
    fig.update_layout(**CHART_CFG, title="Monthly Generation (MWh)",
                      xaxis_title="Month", yaxis_title="MWh")
    return fig

def chart_fleet_gen(all_plants):
    """Grouped bar — fleet total generation per plant."""
    names, gens, prs, avs = [], [], [], []
    for pname, p in all_plants.items():
        k = compute_kpis(p)
        names.append(k["plant_name"])
        gens.append(k["total_gen_mwh"])
        prs.append(k["avg_pr"])
        avs.append(k["availability"])

    fig = go.Figure()
    fig.add_bar(y=names, x=gens, orientation="h", name="Total Gen (MWh)",
                marker_color=PLANT_COLS[:len(names)], opacity=0.85,
                hovertemplate="%{y}: %{x:,.0f} MWh<extra></extra>")
    fig.update_layout(**{**CHART_CFG, "margin": dict(l=110, r=20, t=48, b=50)},
                      title="Fleet — Total Generation (MWh)", xaxis_title="MWh")
    return fig

def chart_fleet_pr(all_plants):
    """PR comparison across all plants."""
    names = [compute_kpis(p)["plant_name"] for p in all_plants.values()]
    prs   = [compute_kpis(p)["avg_pr"]     for p in all_plants.values()]

    fig = go.Figure()
    fig.add_bar(x=names, y=prs, marker_color=[_pc(p) for p in prs], opacity=0.85,
                name="Avg PR %", hovertemplate="%{x}: %{y:.1f}%<extra></extra>")
    fig.add_hline(y=PR_TARGET, line_dash="dash", line_color=RED, line_width=2,
                  annotation_text=f"Target {PR_TARGET}%", annotation_position="right")
    fig.update_layout(**CHART_CFG, title="Plant-wise PR Comparison",
                      xaxis_title="Plant", yaxis_title="PR (%)",
                      yaxis=dict(range=[60, 100]))
    return fig

def chart_inverter(plant, last=30):
    inv_recs = plant.get("inv_recs", [])
    if not inv_recs:
        return _ef("No inverter data in Data_Entry sheet")
    keys = sorted(k for k in inv_recs[0] if k.startswith("Inv_"))
    if not keys:
        return _ef("No Inv_ columns found")

    recent  = inv_recs[-last:]
    totals  = {k: sum(r.get(k,0) for r in recent)/1000 for k in keys}  # MWh
    avg_mwh = sum(totals.values())/len(totals) if totals else 0

    pal = [BLU,GRN,AMB,RED,PUR,TEL,"#f59e0b","#06b6d4","#ec4899","#84cc16"]*3
    fig = go.Figure()
    fig.add_bar(x=list(totals), y=list(totals.values()), name="MWh",
                marker_color=pal[:len(keys)],
                hovertemplate="%{x}: %{y:.2f} MWh<extra></extra>")
    fig.add_hline(y=avg_mwh, line_dash="dash", line_color=GRN,
                  annotation_text=f"Avg {avg_mwh:.2f} MWh", annotation_position="right")
    fig.update_layout(**CHART_CFG, title=f"Inverter Comparison — Last {last} Days (MWh)",
                      xaxis_title="Inverter", yaxis_title="MWh")
    return fig

def chart_downtime(plant):
    events = plant.get("faults", [])
    if not events:
        f = go.Figure()
        f.update_layout(**CHART_CFG, title="Plant Downtime — No Events")
        f.add_annotation(text="No downtime events  ✅", x=0.5, y=0.5,
                         showarrow=False, font=dict(size=16, color="#22c55e"))
        return f

    reason_hrs = defaultdict(float)
    for e in events:
        reason_hrs[e["reason"]] += e["downtime_hrs"]

    sorted_r = sorted(reason_hrs.items(), key=lambda x: -x[1])
    labels  = [r[0] for r in sorted_r]
    hrs     = [r[1] for r in sorted_r]
    total   = sum(hrs) or 1
    cumul   = [sum(hrs[:i+1])/total*100 for i in range(len(hrs))]

    fig = go.Figure()
    fig.add_bar(x=labels, y=hrs, name="Downtime (hrs)", marker_color=RED, opacity=0.8,
                hovertemplate="%{x}: %{y:.1f} hrs<extra></extra>")
    fig.add_scatter(x=labels, y=cumul, name="Cumulative %", yaxis="y2",
                    mode="lines+markers", line=dict(color=AMB, width=2.5),
                    marker=dict(size=7), hovertemplate="%{y:.0f}%<extra></extra>")
    fig.update_layout(**CHART_CFG, title="Downtime Pareto — Top Failure Reasons",
                      xaxis=dict(tickangle=-30), xaxis_title="Reason",
                      yaxis=dict(title="Hours"),
                      yaxis2=dict(title="Cumulative %", overlaying="y", side="right",
                                  range=[0,110], showgrid=False))
    return fig

def chart_avail_gauge(plant):
    avs = plant.get("avail_recs", [])
    if avs:
        vals = [r["availability"] for r in avs if r["availability"] > 0]
        avail = sum(vals)/len(vals) if vals else 100
    else:
        faults  = plant.get("faults",[])
        tot_out = sum(f["downtime_hrs"] for f in faults)
        n = len(plant.get("records",[1]))
        avail = max(0, 100 - tot_out/(n*24)*100)

    color = GRN if avail >= 98 else AMB if avail >= 92 else RED
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=round(avail, 1),
        number={"suffix":"%","font":{"size":34,"color":color}},
        title={"text":"Plant Availability","font":{"size":14}},
        gauge={
            "axis":{"range":[0,100],"ticksuffix":"%"},
            "bar": {"color":color,"thickness":0.28},
            "steps":[
                {"range":[0,92],  "color":"rgba(220,38,38,0.1)"},
                {"range":[92,98], "color":"rgba(217,119,6,0.1)"},
                {"range":[98,100],"color":"rgba(22,163,74,0.1)"},
            ],
            "threshold":{"line":{"color":GRN,"width":3},"thickness":0.8,"value":98},
        },
    ))
    fig.add_annotation(text="Target: ≥ 98%", x=0.5, y=-0.02,
                       showarrow=False, font=dict(size=11, color="#64748b"))
    fig.update_layout(paper_bgcolor="#fff", font=dict(family="Arial"),
                      margin=dict(l=20, r=20, t=50, b=30))
    return fig

def chart_gti_scatter(plant):
    recs = plant["records"]
    gtis = [r["gti"] for r in recs if r["gti"] > 0]
    gens = [r["gen_kwh"]/1000 for r in recs if r["gti"] > 0]
    prs  = [r["pr"] for r in recs if r["gti"] > 0]

    if not gtis:
        return _ef("No GTI data in Budgeted_Data sheet")

    fig = go.Figure(go.Scatter(
        x=gtis, y=gens, mode="markers",
        marker=dict(color=prs, colorscale="RdYlGn", cmin=70, cmax=92,
                    size=8, opacity=0.75,
                    colorbar=dict(title="PR %", thickness=14, len=0.8)),
        hovertemplate="GTI: %{x:.2f} kWh/m²<br>Gen: %{y:.2f} MWh<extra></extra>",
    ))
    fig.update_layout(**CHART_CFG, title="GTI vs Generation  (color = PR %)",
                      xaxis_title="GTI (kWh/m²)", yaxis_title="Daily Generation (MWh)")
    return fig


# ── Global in-memory data store ───────────────────────────────────────────────
# Data lives in Python memory (no JSON serialization overhead).
# dcc.Store only holds a small version counter to trigger re-renders.
_PLANTS = {}   # global dict: plant_name -> plant_data

def _build_opts():
    base = [{"label": "All Plants (Fleet)", "value": "__fleet__"}]
    return base + [{"label": p, "value": p} for p in _PLANTS]

# Auto-load at startup
print("Loading DGR files...")
_d, _e = load_dgr_folder()
if _d:
    _PLANTS.update(_d)
    print(f"Loaded {len(_PLANTS)} plants: {', '.join(_PLANTS.keys())}")
    if _e:  # partial errors
        print(f"  Warnings: {_e}")
else:
    print(f"Auto-load failed: {_e}")

# ── Layout ────────────────────────────────────────────────────────────────────

app = dash.Dash(__name__, title="DGR Dashboard")
server = app.server   # expose Flask server for gunicorn (Render)
EF = _ef()

app.layout = html.Div([
    dcc.Store(id="store", data=len(_PLANTS)),  # version counter only
    dcc.Interval(id="init-tick", interval=400, n_intervals=0, max_intervals=1),

    # Header
    html.Div([
        html.Div([
            html.Div("☀", style={"fontSize":"26px","marginRight":"10px"}),
            html.Div([
                html.H1("DGR Performance Dashboard — FY 2025-26", style={
                    "margin":"0","fontSize":"20px","fontWeight":"700","color":"#ffffff"}),
                html.Div(id="hdr-info", children="Load DGR files to begin",
                         style={"fontSize":"12px","color":"#94a3b8","marginTop":"2px"}),
            ]),
        ], style={"display":"flex","alignItems":"center"}),
        html.Div(id="hdr-grade", style={"marginLeft":"auto"}),
    ], style={
        "background":"linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%)",
        "padding":"14px 28px","display":"flex","alignItems":"center",
        "boxShadow":"0 3px 10px rgba(0,0,0,0.35)",
    }),

    html.Div([

        # Upload row
        html.Div([
            html.Div([
                dcc.Upload(id="upload", multiple=True,
                    children=html.Div([
                        html.Div("📂", style={"fontSize":"28px","marginBottom":"6px"}),
                        html.Div([
                            html.Strong("Drag & drop all DGR Excel files here"),
                            html.Span("  (select multiple files) or "),
                            html.Span("click to browse", style={"color":BLU,"fontWeight":"600"}),
                        ], style={"fontSize":"14px","color":"#475569"}),
                        html.Div("DGR_Devki_Galol_*.xlsx  |  DGR_Kanza_*.xlsx  |  DGR_Mandapara_*.xlsx  |  DGR_Mandodara_*.xlsx",
                                 style={"fontSize":"11px","color":"#94a3b8","marginTop":"4px"}),
                    ], style={"textAlign":"center"}),
                    style={"width":"100%","padding":"22px 16px",
                           "borderWidth":"2px","borderStyle":"dashed","borderColor":"#cbd5e1",
                           "borderRadius":"10px","backgroundColor":"#f8fafc","cursor":"pointer"},
                ),
            ], style={"flex":"1"}),

            html.Div("OR", style={"color":"#94a3b8","fontWeight":"700","margin":"0 16px",
                                   "display":"flex","alignItems":"center"}),

            html.Button([
                html.Div("📁", style={"fontSize":"22px"}),
                html.Div("Load DGR Folder", style={"fontWeight":"700","fontSize":"13px","marginTop":"4px"}),
                html.Div("Auto-scan all 4 plants", style={"fontSize":"11px","opacity":"0.7"}),
            ], id="load-btn", n_clicks=0, style={
                "backgroundColor":"#0f172a","color":"#fff",
                "border":"2px solid #334155","borderRadius":"10px",
                "padding":"14px 20px","cursor":"pointer",
                "textAlign":"center","lineHeight":"1.4","minWidth":"155px",
            }),
        ], style={
            "display":"flex","alignItems":"center","backgroundColor":"#fff",
            "borderRadius":"12px","padding":"18px",
            "boxShadow":"0 1px 4px rgba(0,0,0,0.08)","marginBottom":"16px",
        }),

        html.Div(id="alert"),

        # Plant selector
        html.Div([
            html.Span("View Plant:", style={"fontWeight":"600","color":"#1e293b","marginRight":"10px","fontSize":"13px"}),
            dcc.RadioItems(id="plant-sel",
                           options=_build_opts(),
                           value="__fleet__", inline=True,
                           inputStyle={"marginRight":"4px"},
                           labelStyle={"marginRight":"16px","fontSize":"13px","color":"#475569"}),
        ], style={
            "backgroundColor":"#fff","borderRadius":"10px","padding":"12px 18px",
            "boxShadow":"0 1px 3px rgba(0,0,0,0.07)","marginBottom":"16px",
            "display":"flex","alignItems":"center","flexWrap":"wrap","gap":"6px",
        }),

        # KPI cards
        html.Div(id="kpis", style={"display":"flex","gap":"12px","flexWrap":"wrap","marginBottom":"16px"}),

        # Row 1: Daily gen + PR
        html.Div([
            _card(dcc.Graph(id="g-daily",   figure=EF, config={"displayModeBar":False},
                            style={"height":"340px"}), {"flex":"2.2"}),
            _card(dcc.Graph(id="g-pr",      figure=EF, config={"displayModeBar":False},
                            style={"height":"340px"}), {"flex":"1.4"}),
        ], style={"display":"flex","gap":"14px","marginBottom":"14px"}),

        # Row 2: Monthly + Fleet gen + Availability gauge
        html.Div([
            _card(dcc.Graph(id="g-monthly", figure=EF, config={"displayModeBar":False},
                            style={"height":"290px"}), {"flex":"1.8"}),
            _card(dcc.Graph(id="g-fleet",   figure=EF, config={"displayModeBar":False},
                            style={"height":"290px"}), {"flex":"1.4"}),
            _card(dcc.Graph(id="g-avail",   figure=EF, config={"displayModeBar":False},
                            style={"height":"290px"}), {"flex":"0.8"}),
        ], style={"display":"flex","gap":"14px","marginBottom":"14px"}),

        # Row 3: Inverter + Downtime pareto
        html.Div([
            _card(dcc.Graph(id="g-inv",      figure=EF, config={"displayModeBar":False},
                            style={"height":"280px"}), {"flex":"1.5"}),
            _card(dcc.Graph(id="g-downtime", figure=EF, config={"displayModeBar":False},
                            style={"height":"280px"}), {"flex":"1.5"}),
        ], style={"display":"flex","gap":"14px","marginBottom":"14px"}),

        # Row 4: GTI scatter (full width)
        _card(dcc.Graph(id="g-gti", figure=EF, config={"displayModeBar":False},
                        style={"height":"280px"}), {"marginBottom":"14px"}),

    ], style={"padding":"20px 28px","backgroundColor":"#f0f4f8","minHeight":"calc(100vh - 64px)"}),

    html.Div([
        html.Span(
            f"☀ DGR Performance Dashboard  |  FY 2025-26  |  "
            f"Tariff: ₹{TARIFF}/kWh  |  PR Target: {PR_TARGET}%  |  CO₂ Factor: {CO2_FACTOR} t/kWh",
            style={"color":"#64748b","fontSize":"12px"},
        ),
    ], style={"padding":"10px 28px","backgroundColor":"#1e293b","textAlign":"center"}),

], style={"fontFamily":"Arial, Helvetica, sans-serif","backgroundColor":"#f0f4f8"})


# ── Callbacks ─────────────────────────────────────────────────────────────────

@app.callback(
    Output("store",    "data"),
    Output("alert",    "children"),
    Output("plant-sel","options"),
    Output("plant-sel","value"),
    Input("upload",    "contents"),
    Input("load-btn",  "n_clicks"),
    State("upload",    "filename"),
    State("store",     "data"),
    prevent_initial_call=True,
)
def load_data(contents_list, n_clicks, filenames, version):
    global _PLANTS
    ctx = dash.callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0]

    if trigger == "load-btn":
        result, err = load_dgr_folder()
        if err:
            return version, _alert(err, "error"), _build_opts(), "__fleet__"
        _PLANTS.clear()
        _PLANTS.update(result)
        banner = _alert(f"Loaded {len(_PLANTS)} plants: {', '.join(_PLANTS.keys())}", "success")

    elif trigger == "upload" and contents_list:
        errors = []
        for contents, fname in zip(contents_list, filenames):
            try:
                _, b64 = contents.split(",", 1)
                data, err = parse_dgr(base64.b64decode(b64), fname)
                if err: errors.append(err)
                else:   _PLANTS[data["plant_name"]] = data
            except Exception as ex:
                errors.append(f"{fname}: {ex}")
        if not _PLANTS:
            return version, _alert("No valid DGR files. " + " | ".join(errors), "error"), _build_opts(), "__fleet__"
        msg = f"Loaded {len(_PLANTS)} plants: {', '.join(_PLANTS.keys())}"
        if errors: msg += f"  |  Warnings: {'; '.join(errors)}"
        banner = _alert(msg, "success")
    else:
        return version, None, _build_opts(), "__fleet__"

    return (version or 0) + 1, banner, _build_opts(), "__fleet__"


@app.callback(
    Output("kpis",      "children"),
    Output("hdr-info",  "children"),
    Output("hdr-grade", "children"),
    Output("g-daily",   "figure"),
    Output("g-pr",      "figure"),
    Output("g-monthly", "figure"),
    Output("g-fleet",   "figure"),
    Output("g-avail",   "figure"),
    Output("g-inv",     "figure"),
    Output("g-downtime","figure"),
    Output("g-gti",     "figure"),
    Input("store",      "data"),
    Input("plant-sel",  "value"),
    Input("init-tick",  "n_intervals"),
    prevent_initial_call=False,   # Dash 4.x: explicitly fire on first page load
)
def update_dash(version, selected, _tick):
    ef = _ef()
    empty_ret = ([], "No data loaded — run the dashboard and it will auto-load", None,
                 ef, ef, ef, ef, ef, ef, ef, ef)
    all_plants = _PLANTS
    if not all_plants: return empty_ret

    fleet_kpis = [compute_kpis(all_plants[p]) for p in all_plants]

    if selected == "__fleet__" or selected not in all_plants:
        # Fleet aggregates
        n_plants  = len(fleet_kpis)
        total_gen = sum(k["total_gen_mwh"]  for k in fleet_kpis)
        avg_pr    = sum(k["avg_pr"]         for k in fleet_kpis) / n_plants
        avg_av    = sum(k["availability"]   for k in fleet_kpis) / n_plants
        avg_cuf   = sum(k["avg_cuf"]        for k in fleet_kpis) / n_plants
        tot_rev   = sum(k["revenue_lakh"]   for k in fleet_kpis)
        tot_co2   = sum(k["co2_tonnes"]     for k in fleet_kpis)
        tot_faults= sum(k["n_faults"]       for k in fleet_kpis)

        grade, gc = (
            ("Excellent",GRN) if avg_pr>=80 else
            ("Good","#22c55e") if avg_pr>=75 else
            ("Fair",AMB) if avg_pr>=65 else ("Poor",RED)
        )
        cards = [
            kpi_card("Fleet Generation", f"{total_gen:,.1f}", "MWh",
                     f"All {n_plants} plants combined\nFY 2025-26", BLU, "⚡"),
            kpi_card("Avg Perf. Ratio", f"{avg_pr:.1f}", "%",
                     f"Target: {PR_TARGET}%\nGrade: {grade}", GRN if avg_pr>=PR_TARGET else AMB, "📊"),
            kpi_card("Avg Availability", f"{avg_av:.1f}", "%",
                     f"Target: ≥98%\nFaults: {tot_faults}", GRN if avg_av>=98 else AMB, "✅"),
            kpi_card("Avg CUF", f"{avg_cuf:.2f}", "%",
                     f"Capacity Utilization\nFleet average", TEL, "🔋"),
            kpi_card("Total Revenue", f"₹{tot_rev:.1f}", "Lakh",
                     f"@ ₹{TARIFF}/kWh\nCO₂ Saved: {tot_co2:.0f} t", GRN, "💰"),
            kpi_card("Total Faults", f"{tot_faults}", "Events",
                     f"All plants combined\nSelect plant for detail", RED if tot_faults>20 else AMB, "⚠️"),
        ]
        hdr_info  = f"{n_plants} Plants | Fleet Overview | FY 2025-26"
        hdr_grade = html.Div(f"Fleet Grade: {grade}", style={
            "backgroundColor":gc,"color":"#fff","padding":"6px 18px",
            "borderRadius":"20px","fontWeight":"700","fontSize":"13px"})

        first = all_plants[list(all_plants.keys())[0]]
        return (cards, hdr_info, hdr_grade,
                chart_daily_gen(first), chart_fleet_pr(all_plants),
                chart_monthly(first),  chart_fleet_gen(all_plants),
                chart_avail_gauge(first),
                chart_inverter(first), chart_downtime(first),
                chart_gti_scatter(first))

    else:
        plant = all_plants[selected]
        k = compute_kpis(plant)
        cards = [
            kpi_card("Total Generation", f"{k['total_gen_mwh']:,.1f}", "MWh",
                     f"DC: {k['dc_kw']} kW  |  {k['n_days']} days\n{k['date_from']} → {k['date_to']}", BLU, "⚡"),
            kpi_card("Performance Ratio", f"{k['avg_pr']}", "%",
                     f"Target: {PR_TARGET}%\nGrade: {k['grade']}", GRN if k['avg_pr']>=PR_TARGET else AMB, "📊"),
            kpi_card("Plant Availability", f"{k['availability']}", "%",
                     f"Outage: {k['outage_hrs']:.1f} hrs\nFaults: {k['n_faults']}", GRN if k['availability']>=98 else AMB, "✅"),
            kpi_card("CUF", f"{k['avg_cuf']:.2f}", "%",
                     f"Capacity: {k['dc_kw']} kW DC", TEL, "🔋"),
            kpi_card("Revenue", f"₹{k['revenue_lakh']:.2f}", "Lakh",
                     f"@ ₹{TARIFF}/kWh\nCO₂: {k['co2_tonnes']:.0f} tonnes", GRN, "💰"),
            kpi_card("Faults", f"{k['n_faults']}", "Events",
                     f"Downtime: {k['outage_hrs']:.1f} hrs", RED if k['n_faults']>10 else AMB, "⚠️"),
        ]
        hdr_info  = f"{selected}  |  DC: {k['dc_kw']} kW  |  {k['date_from']} → {k['date_to']}  ({k['n_days']} days)"
        hdr_grade = html.Div(f"{k['grade']}  ▸  PR {k['avg_pr']}%", style={
            "backgroundColor":k["grade_color"],"color":"#fff","padding":"6px 18px",
            "borderRadius":"20px","fontWeight":"700","fontSize":"13px"})

        return (cards, hdr_info, hdr_grade,
                chart_daily_gen(plant), chart_pr_trend(plant),
                chart_monthly(plant),   chart_fleet_gen(all_plants),
                chart_avail_gauge(plant),
                chart_inverter(plant),  chart_downtime(plant),
                chart_gti_scatter(plant))


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print()
    print("=" * 68)
    print("  ☀  DGR Performance Dashboard — FY 2025-26")
    print("=" * 68)
    print(f"  DGR Folder : {DGR_FOLDER}")
    print(f"  Tariff     : ₹{TARIFF}/kWh  |  PR Target: {PR_TARGET}%")
    print("=" * 68)
    print("  Open browser: http://localhost:8051")
    print("=" * 68)
    print()
    port = int(os.environ.get("PORT", 8051))
    app.run(debug=False, port=port, host="0.0.0.0")
